#!/usr/bin/env python3
"""
Reddit VOC 原始帖子与评论采集脚本
- 拉取指定母婴 subreddits 最近半年（180 天）内的帖子。数据来源：PRAW（new + 关键词 search + 时间分片）
  + 可选 PullPush（--pullpush 或 USE_PULLPUSH=1）。并抓取每帖下方评论（仅 PRAW 来源有评论），写入 Markdown。
- 依赖：praw；启用 PullPush 时需 requests。见 scripts/README.md 配置环境变量后运行。
"""

import argparse
import os
import re
import time
from datetime import datetime, timedelta
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
VOC_DIR = SCRIPT_DIR.parent
RAW_LOGS = VOC_DIR / "logs" / "raw"

try:
    from dotenv import load_dotenv
    load_dotenv(SCRIPT_DIR / ".env")
except ImportError:
    pass

try:
    import praw
except ImportError:
    print("请先安装: pip install praw")
    raise

try:
    import requests
except ImportError:
    requests = None

try:
    from openpyxl import Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

PULLPUSH_API = "https://api.pullpush.io/reddit/search/submission/"

SUBREDDITS = [
    "breastfeeding",
    "WorkingMoms",
    "ExclusivelyPumping",
    "BabyBumps",
    "beyondthebump",
    "NewParents",
]

# 最近半年（天）
ROLLING_DAYS = 180

# 时间分片：将 180 天按多少天切段，每段用 CloudSearch timestamp 拉取（每段最多 1000 条）
CHUNK_DAYS = 10

# 每 subreddit 从 new 列表拉取的最大条数（Reddit 单次 listing 上限约 1000）
LIMIT_NEW_PER_SUB = 1000

# 每 subreddit 按关键词 search、time_filter=year 的条数，用于补全半年内相关帖
SEARCH_LIMIT_PER_SUB = 300

# 关键词（用于 search 补全半年内与吸奶器/背奶/品牌相关帖）
SEARCH_QUERIES = [
    "pump",
    "wearable pump",
    "breast pump",
    "Momcozy",
    "pumping at work",
    "back to work",
]

# 第二轮及以后使用的扩展关键词（多轮采集时用）
SEARCH_QUERIES_EXTENDED = [
    "Spectra",
    "Elvie",
    "Willow",
    "Medela",
    "leakage",
    "quiet pump",
    "hands free",
    "insurance",
    "FSA",
    "return to work",
]

# 多轮采集：共跑几轮（每轮包含 new + search，轮次间用不同关键词或 hot 补全）
NUM_COLLECTION_ROUNDS = 3

# 每轮 search 单关键词最大条数
SEARCH_LIMIT_PER_QUERY = 100

# 每帖最多保留的评论数（仅顶层评论，避免单文件过大）
MAX_COMMENTS_PER_POST = 50


def sanitize(s: str, max_len: int = 400) -> str:
    if not s:
        return ""
    s = re.sub(r"\s+", " ", s.strip())
    return s[:max_len] + ("..." if len(s) > max_len else "")


def fetch_posts_in_window(reddit, cutoff_ts: float, round_index: int = 0):
    """
    拉取各 sub 的 new 列表 + 关键词 search，合并去重，只保留 created_utc >= cutoff_ts 的帖子。
    round_index: 第几轮（0-based）。第 0 轮用 new + SEARCH_QUERIES；第 1 轮用 hot + SEARCH_QUERIES_EXTENDED；第 2 轮再用 new 小量 + 扩展关键词补漏。
    """
    seen_ids = set()
    posts_by_id = {}
    queries = SEARCH_QUERIES if round_index == 0 else (SEARCH_QUERIES_EXTENDED if round_index == 1 else SEARCH_QUERIES + SEARCH_QUERIES_EXTENDED)
    limit_new = LIMIT_NEW_PER_SUB if round_index == 0 else (300 if round_index == 2 else 0)
    limit_hot = 300 if round_index == 1 else 0

    for sub_name in SUBREDDITS:
        try:
            sub = reddit.subreddit(sub_name)
            if limit_new:
                for post in sub.new(limit=limit_new):
                    if post.id in seen_ids or post.created_utc < cutoff_ts:
                        continue
                    seen_ids.add(post.id)
                    posts_by_id[post.id] = post
                time.sleep(0.3)
            if limit_hot:
                for post in sub.hot(limit=limit_hot):
                    if post.id in seen_ids or post.created_utc < cutoff_ts:
                        continue
                    seen_ids.add(post.id)
                    posts_by_id[post.id] = post
                time.sleep(0.3)
            for q in queries:
                try:
                    for post in sub.search(q, time_filter="year", limit=min(SEARCH_LIMIT_PER_QUERY, 100)):
                        if post.id in seen_ids or post.created_utc < cutoff_ts:
                            continue
                        seen_ids.add(post.id)
                        posts_by_id[post.id] = post
                except Exception:
                    pass
                time.sleep(0.5)
            time.sleep(0.5)
        except Exception as e:
            print(f"Warning: {sub_name} -> {e}")

    return list(posts_by_id.values())


def fetch_posts_praw_timestamp_chunks(reddit, cutoff_ts: float):
    """
    按时间分片拉取各 sub 的帖子，突破单次 1000 条上限：每段用 CloudSearch timestamp:start..end，
    每段最多 1000 条，用本段最早一条的 created_utc 作为下一段 end 分页，直到覆盖 cutoff_ts 或取满。
    """
    posts_by_id = {}
    chunk_seconds = CHUNK_DAYS * 86400
    now_ts = time.time()

    for sub_name in SUBREDDITS:
        try:
            sub = reddit.subreddit(sub_name)
            seg_end = now_ts
            while seg_end > cutoff_ts:
                seg_start = max(cutoff_ts, seg_end - chunk_seconds)
                try:
                    q = f"timestamp:{int(seg_start)}..{int(seg_end)}"
                    batch = list(sub.search(q, sort="new", limit=None, syntax="cloudsearch"))
                except Exception as e:
                    print(f"  PRAW timestamp chunk {sub_name} {seg_start}-{seg_end}: {e}")
                    break
                for post in batch:
                    if post.created_utc < cutoff_ts:
                        continue
                    posts_by_id[post.id] = post
                if len(batch) < 1000:
                    seg_end = seg_start - 1
                else:
                    seg_end = min(p.created_utc for p in batch)
                time.sleep(0.5)
            time.sleep(0.3)
        except Exception as e:
            print(f"Warning: {sub_name} (timestamp chunks) -> {e}")

    return list(posts_by_id.values())


def fetch_posts_pullpush(cutoff_ts: float):
    """
    从 PullPush API 按 subreddit + after 拉取帖子，size=500 分页，before=本批最小 created_utc 直到无数据。
    返回 list[dict]，每个 dict 含 id, subreddit, title, url, created_utc, selftext, num_comments, score（与表格列一致）。
    """
    if requests is None:
        print("PullPush 需要 requests，请安装: pip install requests")
        return []
    posts_by_id = {}
    for sub_name in SUBREDDITS:
        before_ts = None
        while True:
            try:
                params = {
                    "subreddit": sub_name,
                    "after": int(cutoff_ts),
                    "size": 500,
                    "sort": "desc",
                    "sort_type": "created_utc",
                }
                if before_ts is not None:
                    params["before"] = before_ts
                r = requests.get(PULLPUSH_API, params=params, timeout=30)
                r.raise_for_status()
                data = r.json()
            except Exception as e:
                print(f"  PullPush {sub_name}: {e}")
                break
            items = data.get("data") if isinstance(data, dict) else []
            if not items:
                break
            for it in items:
                created = it.get("created_utc") or 0
                if created < cutoff_ts:
                    continue
                pid = it.get("id")
                if not pid:
                    continue
                permalink = it.get("permalink") or ""
                if not permalink.startswith("http"):
                    permalink = "https://old.reddit.com" + (permalink if permalink.startswith("/") else "/" + permalink)
                posts_by_id[pid] = {
                    "id": pid,
                    "subreddit": it.get("subreddit") or sub_name,
                    "title": (it.get("title") or "").strip(),
                    "url": permalink,
                    "created_utc": created,
                    "selftext": (it.get("selftext") or "").strip(),
                    "num_comments": int(it.get("num_comments") or 0),
                    "score": int(it.get("score") or 0),
                }
            if len(items) < 500:
                break
            before_ts = min(it.get("created_utc") or 0 for it in items)
            if before_ts <= int(cutoff_ts):
                break
            time.sleep(0.5)
        time.sleep(0.3)
    return list(posts_by_id.values())


def fetch_comments_for_post(post, reddit, max_comments: int = MAX_COMMENTS_PER_POST):
    """抓取帖子下顶层评论，最多 max_comments 条。返回 [(body, created_utc, score, permalink), ...]"""
    out = []
    try:
        post.comments.replace_more(limit=0)
        for i, c in enumerate(post.comments):
            if i >= max_comments:
                break
            if getattr(c, "body", None) is None:
                continue
            body = (c.body or "").strip()
            if not body:
                continue
            created = getattr(c, "created_utc", 0)
            score = getattr(c, "score", 0)
            permalink = "https://old.reddit.com" + (getattr(c, "permalink", None) or "")
            out.append((body, created, score, permalink))
    except Exception as e:
        pass
    return out


def main():
    parser = argparse.ArgumentParser(description="Reddit VOC 帖子与评论采集（PRAW 时间分片 + 可选 PullPush）")
    parser.add_argument("--pullpush", action="store_true", help="同时从 PullPush API 拉取帖子并合并（也可用 USE_PULLPUSH=1）")
    args = parser.parse_args()
    use_pullpush = args.pullpush or (os.environ.get("USE_PULLPUSH", "").strip().lower() in ("1", "true", "yes"))

    client_id = os.environ.get("REDDIT_CLIENT_ID")
    client_secret = os.environ.get("REDDIT_CLIENT_SECRET")
    username = os.environ.get("REDDIT_USERNAME")
    password = os.environ.get("REDDIT_PASSWORD")

    has_reddit = all([client_id, client_secret, username, password])
    if not has_reddit and (args.pullpush or os.environ.get("USE_PULLPUSH", "").strip().lower() in ("1", "true", "yes")):
        # 无 Reddit 凭证时：仅用 PullPush 拉取帖子（无评论），仍可产出 raw + Excel
        if requests is None:
            print("PullPush 需 requests，请安装: pip install requests")
            return 1
        print("未设置 Reddit API 凭证，仅使用 PullPush 拉取帖子（无评论）。配置凭证后可启用 PRAW + 评论，见 README。")
        cutoff = (datetime.utcnow() - timedelta(days=ROLLING_DAYS)).timestamp()
        pullpush_list = fetch_posts_pullpush(cutoff)
        posts = list({d["id"]: d for d in pullpush_list}.values())
        posts.sort(key=lambda p: p["created_utc"], reverse=True)
        RAW_LOGS.mkdir(parents=True, exist_ok=True)
        today = datetime.now().strftime("%Y-%m-%d")
        out_path = RAW_LOGS / f"{today}-reddit-posts.md"
        rows = []
        for post in posts:
            created = datetime.utcfromtimestamp(post["created_utc"]).strftime("%Y-%m-%d %H:%M UTC")
            selftext = sanitize(post.get("selftext") or "")
            title = sanitize(post.get("title") or "", max_len=200)
            url = post.get("url") or ""
            sub_name = post.get("subreddit") or ""
            rows.append({
                "id": post["id"],
                "subreddit": f"r/{sub_name}" if sub_name and not sub_name.startswith("r/") else sub_name or "r/?",
                "title": title,
                "url": url,
                "created": created,
                "selftext": selftext,
                "num_comments": post.get("num_comments", 0),
                "score": post.get("score", 0),
                "post_obj": None,
            })
        rows.sort(key=lambda x: x["created"], reverse=True)
        sources_note = "PullPush（仅帖子，无 Reddit 凭证故无评论）"
        with open(out_path, "w", encoding="utf-8") as f:
            f.write(f"# Reddit 原始帖子与评论 {today}\n\n")
            f.write(f"> 由 `reddit_voc_fetch.py` 生成。数据来源：**{sources_note}**。范围：**最近 {ROLLING_DAYS} 天**。共 **{len(rows)}** 条帖子（无评论）。\n\n")
            f.write("---\n\n## 帖子列表\n\n")
            f.write("| 序号 | subreddit | title | url | created | num_comments | score | selftext（摘要） |\n")
            f.write("|------|-----------|-------|-----|---------|--------------|-------|------------------|\n")
            for i, r in enumerate(rows, 1):
                selftext_cell = (r["selftext"] or "").replace("|", "\\|").replace("\n", " ")[:300]
                if len((r["selftext"] or "")) > 300:
                    selftext_cell += "..."
                title_cell = (r["title"] or "").replace("|", "\\|").replace("\n", " ")
                f.write(f"| {i} | {r['subreddit']} | {title_cell} | {r['url']} | {r['created']} | {r['num_comments']} | {r['score']} | {selftext_cell} |\n")
            f.write("\n---\n\n## 帖子正文与评论\n\n")
            for i, r in enumerate(rows, 1):
                f.write(f"### [{i}] {r['subreddit']} | {r['created']} | [原帖链接]({r['url']})\n\n")
                f.write(f"**标题**：{r['title']}\n\n")
                if r["selftext"]:
                    f.write(f"**正文**：\n\n{r['selftext']}\n\n")
                f.write("**评论**：PullPush 归档，无实时评论。\n\n\n")
        print(f"已写入 {len(rows)} 条帖子（无评论）-> {out_path}")
        if OPENPYXL_AVAILABLE:
            out_xlsx = RAW_LOGS / f"{today}-reddit-posts.xlsx"
            wb = Workbook()
            ws_posts = wb.active
            ws_posts.title = "帖子"
            ws_posts.append(["post_id", "post_url", "subreddit", "title", "created", "selftext", "num_comments", "score"])
            for r in rows:
                ws_posts.append([r["id"], r["url"], r["subreddit"], r["title"], r["created"], (r["selftext"] or "")[:32767], r["num_comments"], r["score"]])
            ws_comments = wb.create_sheet("评论", 1)
            ws_comments.append(["post_id", "post_url", "comment_index", "comment_body", "comment_created_readable", "comment_score", "comment_link"])
            wb.save(out_xlsx)
            print(f"已写入 Excel（仅帖子，含链接）-> {out_xlsx}")
        return 0

    if not has_reddit:
        print("请设置环境变量: REDDIT_CLIENT_ID, REDDIT_CLIENT_SECRET, REDDIT_USERNAME, REDDIT_PASSWORD")
        print("获取方式见: books/maternal_social_voc/scripts/README.md")
        return 1

    reddit = praw.Reddit(
        client_id=client_id,
        client_secret=client_secret,
        username=username,
        password=password,
        user_agent="voc_maternal_brand_fetch/1.0 (personal use; maternal VOC research)",
    )
    reddit.read_only = True

    cutoff = (datetime.utcnow() - timedelta(days=ROLLING_DAYS)).timestamp()
    print(f"拉取最近 {ROLLING_DAYS} 天内的帖子（cutoff = {datetime.utcfromtimestamp(cutoff).strftime('%Y-%m-%d')} UTC），共 {NUM_COLLECTION_ROUNDS} 轮采集...")

    all_posts_by_id = {}
    for r in range(NUM_COLLECTION_ROUNDS):
        round_posts = fetch_posts_in_window(reddit, cutoff, round_index=r)
        for p in round_posts:
            all_posts_by_id[p.id] = p
        print(f"  第 {r+1} 轮后累计帖子数: {len(all_posts_by_id)}")
        time.sleep(1)
    chunk_posts = fetch_posts_praw_timestamp_chunks(reddit, cutoff)
    for p in chunk_posts:
        all_posts_by_id[p.id] = p
    print(f"  PRAW 时间分片后累计帖子数: {len(all_posts_by_id)}")
    if use_pullpush:
        pullpush_list = fetch_posts_pullpush(cutoff)
        for d in pullpush_list:
            if d["id"] not in all_posts_by_id:
                all_posts_by_id[d["id"]] = d
        print(f"  PullPush 合并后累计帖子数: {len(all_posts_by_id)}")
    posts = list(all_posts_by_id.values())
    posts.sort(key=lambda p: (p.created_utc if hasattr(p, "created_utc") else p["created_utc"]), reverse=True)

    RAW_LOGS.mkdir(parents=True, exist_ok=True)
    today = datetime.now().strftime("%Y-%m-%d")
    out_path = RAW_LOGS / f"{today}-reddit-posts.md"

    # 先写入帖子表格（post 可能为 PRAW submission 或 PullPush 的 dict）
    rows = []
    for post in posts:
        if isinstance(post, dict):
            created = datetime.utcfromtimestamp(post["created_utc"]).strftime("%Y-%m-%d %H:%M UTC")
            selftext = sanitize(post.get("selftext") or "")
            title = sanitize(post.get("title") or "", max_len=200)
            url = post.get("url") or ""
            sub_name = post.get("subreddit") or ""
            rows.append({
                "id": post["id"],
                "subreddit": f"r/{sub_name}" if sub_name and not sub_name.startswith("r/") else sub_name or "r/?",
                "title": title,
                "url": url,
                "created": created,
                "selftext": selftext,
                "num_comments": post.get("num_comments", 0),
                "score": post.get("score", 0),
                "post_obj": None,
            })
        else:
            created = datetime.utcfromtimestamp(post.created_utc).strftime("%Y-%m-%d %H:%M UTC")
            selftext = sanitize(getattr(post, "selftext", "") or "")
            title = sanitize(post.title, max_len=200)
            url = f"https://old.reddit.com{post.permalink}"
            sub_name = post.subreddit.display_name
            rows.append({
                "id": post.id,
                "subreddit": f"r/{sub_name}",
                "title": title,
                "url": url,
                "created": created,
                "selftext": selftext,
                "num_comments": getattr(post, "num_comments", 0),
                "score": getattr(post, "score", 0),
                "post_obj": post,
            })
    rows.sort(key=lambda x: x["created"], reverse=True)

    sources_note = "PRAW（时间分片）" + (" + PullPush" if use_pullpush else "")
    all_comments = []  # 用于 Excel：(post_id, post_url, comment_index, body, created_utc, score, comment_link)

    with open(out_path, "w", encoding="utf-8") as f:
        f.write(f"# Reddit 原始帖子与评论 {today}\n\n")
        f.write(f"> 由 `reddit_voc_fetch.py` 生成。数据来源：**{sources_note}**。范围：**最近 {ROLLING_DAYS} 天**（滚动半年）。共 **{len(rows)}** 条帖子，每条含正文及下方评论（PRAW 来源有评论，PullPush 归档无实时评论）。请从中筛选高价值帖/评论，复制到当日 VOC 表并打标签。\n\n")
        f.write("---\n\n## 帖子列表\n\n")
        f.write("| 序号 | subreddit | title | url | created | num_comments | score | selftext（摘要） |\n")
        f.write("|------|-----------|-------|-----|---------|--------------|-------|------------------|\n")
        for i, r in enumerate(rows, 1):
            selftext_cell = (r["selftext"] or "").replace("|", "\\|").replace("\n", " ")
            if len(selftext_cell) > 300:
                selftext_cell = selftext_cell[:300] + "..."
            title_cell = (r["title"] or "").replace("|", "\\|").replace("\n", " ")
            f.write(f"| {i} | {r['subreddit']} | {title_cell} | {r['url']} | {r['created']} | {r['num_comments']} | {r['score']} | {selftext_cell} |\n")

        f.write("\n---\n\n## 帖子正文与评论\n\n")
        for i, r in enumerate(rows, 1):
            post = r["post_obj"]
            f.write(f"### [{i}] {r['subreddit']} | {r['created']} | [原帖链接]({r['url']})\n\n")
            f.write(f"**标题**：{r['title']}\n\n")
            if r["selftext"]:
                f.write(f"**正文**：\n\n{r['selftext']}\n\n")
            if post is not None:
                comments = fetch_comments_for_post(post, reddit)
                f.write(f"**评论（共 {len(comments)} 条，仅保留顶层）**：\n\n")
                for j, (body, created_utc, score, permalink) in enumerate(comments[:MAX_COMMENTS_PER_POST], 1):
                    created_str = datetime.utcfromtimestamp(created_utc).strftime("%Y-%m-%d %H:%M UTC")
                    body_esc = (body or "").replace("\n", " ").strip()[:500]
                    if len((body or "")) > 500:
                        body_esc += "..."
                    f.write(f"- [{created_str}] (score={score}) [评论链接]({permalink})\n  {body_esc}\n\n")
                    all_comments.append({
                        "post_id": r["id"],
                        "post_url": r["url"],
                        "comment_index": j,
                        "body": body or "",
                        "created_utc": created_utc,
                        "score": score,
                        "comment_link": permalink,
                    })
            else:
                f.write("**评论**：PullPush 归档，无实时评论。\n\n")
            f.write("\n")
            if post is not None:
                time.sleep(0.2)

    print(f"已写入 {len(rows)} 条帖子及评论 -> {out_path}")

    # 同时写入 Excel：帖子表（含链接）+ 评论表
    if OPENPYXL_AVAILABLE:
        out_xlsx = RAW_LOGS / f"{today}-reddit-posts.xlsx"
        wb = Workbook()
        # 工作表1：帖子
        ws_posts = wb.active
        ws_posts.title = "帖子"
        posts_headers = ["post_id", "post_url", "subreddit", "title", "created", "selftext", "num_comments", "score"]
        ws_posts.append(posts_headers)
        for r in rows:
            ws_posts.append([
                r["id"],
                r["url"],
                r["subreddit"],
                r["title"],
                r["created"],
                (r["selftext"] or "")[:32767],
                r["num_comments"],
                r["score"],
            ])
        # 工作表2：评论
        ws_comments = wb.create_sheet("评论", 1)
        comment_headers = ["post_id", "post_url", "comment_index", "comment_body", "comment_created_utc", "comment_created_readable", "comment_score", "comment_link"]
        ws_comments.append(comment_headers)
        for c in all_comments:
            ws_comments.append([
                c["post_id"],
                c["post_url"],
                c["comment_index"],
                (c["body"] or "")[:32767],
                c["created_utc"],
                datetime.utcfromtimestamp(c["created_utc"]).strftime("%Y-%m-%d %H:%M UTC"),
                c["score"],
                c["comment_link"],
            ])
        wb.save(out_xlsx)
        print(f"已写入 Excel（帖子+评论，含链接）-> {out_xlsx}")
    else:
        print("提示：安装 openpyxl 后可同时导出 Excel（pip install openpyxl）")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
